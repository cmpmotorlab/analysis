"""
Walton Compressor Calorimeter Comparison Dashboard
Flask backend: ingest a folder of PDF calorimeter reports (organized as
MainFolder/ID-xxxx.xx/*.pdf), extract key parameters, and present RPM-wise
comparison tables across compressor IDs, with XLSX / PDF export.
"""
import io
import os
import re
import statistics
from collections import defaultdict

from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename

from parser import parse_pdf_bytes, PARAM_META, PARAM_ORDER

# The three headline parameters tracked with range + std-dev at the top of every RPM block
HEADLINE_PARAMS = ['cop', 'cooling_capacity', 'input_power']

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200 MB

# Simple in-memory store (single-user local tool)
STORE = {
    'rows': []   # list of parsed row dicts
}


def _folder_id_from_path(path):
    """Extract the compressor ID folder name from a relative path like
    'MainFolder/ID-7338.01/report.pdf' -> 'ID-7338.01'"""
    parts = [p for p in re.split(r'[\\/]', path) if p]
    if len(parts) >= 2:
        return parts[-2]
    return None


@app.route('/')
def index():
    return render_template('index.html', param_meta=PARAM_META, param_order=PARAM_ORDER)


@app.route('/upload', methods=['POST'])
def upload():
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'No files received'}), 400

    rows = []
    errors = []
    for f in files:
        filename = f.filename or ""
        if not filename.lower().endswith('.pdf'):
            continue
        try:
            data = f.read()
            folder_id = _folder_id_from_path(filename)
            row = parse_pdf_bytes(data, fallback_id=folder_id, source_name=os.path.basename(filename))
            rows.append(row)
        except Exception as e:
            errors.append({'file': filename, 'error': str(e)})

    STORE['rows'] = rows

    return jsonify({
        'count': len(rows),
        'errors': errors,
        'data': build_comparison_payload(rows)
    })


def compute_headline_stats(id_rows):
    """For a single RPM-tier block, compute min/max/avg/std across compressor
    IDs for the headline parameters (COP, Cooling Capacity, Input Power)."""
    stats = {}
    for pkey in HEADLINE_PARAMS:
        vals = [r[pkey] for r in id_rows.values() if r.get(pkey) is not None]
        if not vals:
            stats[pkey] = None
            continue
        stats[pkey] = {
            'min': min(vals),
            'max': max(vals),
            'avg': sum(vals) / len(vals),
            'std': statistics.stdev(vals) if len(vals) > 1 else 0.0,
            'n': len(vals),
        }
    return stats


def build_comparison_payload(rows):
    """Group parsed rows by rpm_tier -> compressor_id -> row, plus summary stats."""
    tiers = defaultdict(dict)
    ids_seen = set()
    for r in rows:
        tier = r.get('rpm_tier')
        sid = r.get('sample_id')
        if tier is None or sid is None:
            continue
        tiers[tier][sid] = r
        ids_seen.add(sid)

    sorted_tiers = sorted(tiers.keys())
    payload_tiers = []
    for t in sorted_tiers:
        id_rows = tiers[t]
        payload_tiers.append({
            'tier': t,
            'ids': sorted(id_rows.keys()),
            'rows': id_rows,
            'stats': compute_headline_stats(id_rows)
        })

    # summary KPIs
    cops = [r['cop'] for r in rows if r.get('cop') is not None]
    summary = {
        'total_reports': len(rows),
        'unique_ids': len(ids_seen),
        'rpm_tiers': len(sorted_tiers),
        'avg_cop': round(sum(cops) / len(cops), 3) if cops else None,
        'max_cop': round(max(cops), 3) if cops else None,
        'min_cop': round(min(cops), 3) if cops else None,
    }

    return {
        'tiers': payload_tiers,
        'ids': sorted(ids_seen),
        'summary': summary
    }


@app.route('/data')
def data():
    return jsonify(build_comparison_payload(STORE['rows']))


@app.route('/export/xlsx', methods=['POST'])
def export_xlsx():
    selected = request.json.get('params', PARAM_ORDER)
    buf = build_xlsx(STORE['rows'], selected)
    return send_file(buf, as_attachment=True,
                      download_name='calorimeter_comparison.xlsx',
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/pdf', methods=['POST'])
def export_pdf():
    selected = request.json.get('params', PARAM_ORDER)
    buf = build_pdf(STORE['rows'], selected)
    return send_file(buf, as_attachment=True,
                      download_name='calorimeter_comparison.pdf',
                      mimetype='application/pdf')


def make_tier_chart_png(tier_block, param_key):
    """Bar chart of a single headline parameter across compressor IDs for one RPM tier."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    ids = tier_block['ids']
    vals = [tier_block['rows'][i].get(param_key) for i in ids]
    label = PARAM_META[param_key]['label']

    fig, ax = plt.subplots(figsize=(5.2, 2.6), dpi=150)
    bar_colors = ['#3d8bfd'] * len(ids)
    if vals and any(v is not None for v in vals):
        clean = [v if v is not None else 0 for v in vals]
        best_idx = clean.index(max(clean))
        bar_colors[best_idx] = '#4caf7d'
    ax.bar(ids, [v if v is not None else 0 for v in vals], color=bar_colors, width=0.55)
    ax.set_title(f"{label} — {tier_block['tier']} RPM", fontsize=10, color='#1F3864', fontweight='bold')
    ax.tick_params(axis='x', labelrotation=40, labelsize=7)
    ax.tick_params(axis='y', labelsize=7)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format='png')
    plt.close(fig)
    buf.seek(0)
    return buf


def build_xlsx(rows, selected_params):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    payload = build_comparison_payload(rows)
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14, color="1F3864")
    stat_label_font = Font(bold=True, size=10, color="1F3864")
    best_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    stat_fill = PatternFill(start_color="EEF3FA", end_color="EEF3FA", fill_type="solid")
    thin = Side(style='thin', color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    for tier_block in payload['tiers']:
        tier = tier_block['tier']
        ids = tier_block['ids']
        id_rows = tier_block['rows']
        stats = tier_block['stats']
        n_data_cols = max(len(ids), 1)
        ws = wb.create_sheet(title=f"{tier} RPM")

        ws.cell(row=1, column=1, value=f"Calorimeter Comparison — {tier} RPM Test Point").font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n_data_cols, 4))

        # ---- Headline stats block (range + std dev for COP / Cooling / Input Power) ----
        stat_header_row = 3
        headers = ["Metric", "Min", "Max", "Avg", "Std Dev"]
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=stat_header_row, column=j, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        r = stat_header_row + 1
        for pkey in HEADLINE_PARAMS:
            s = stats.get(pkey)
            label = PARAM_META[pkey]['label']
            ws.cell(row=r, column=1, value=label).font = stat_label_font
            ws.cell(row=r, column=1).fill = stat_fill
            if s:
                ws.cell(row=r, column=2, value=round(s['min'], 3))
                ws.cell(row=r, column=3, value=round(s['max'], 3))
                ws.cell(row=r, column=4, value=round(s['avg'], 3))
                ws.cell(row=r, column=5, value=round(s['std'], 3))
            for j in range(1, 6):
                ws.cell(row=r, column=j).border = border
                if j > 1:
                    ws.cell(row=r, column=j).alignment = center
            r += 1

        # ---- Main transposed comparison table: Parameter rows x Compressor ID columns ----
        table_top = r + 2
        ws.cell(row=table_top, column=1, value="Parameter").font = header_font
        ws.cell(row=table_top, column=1).fill = header_fill
        ws.cell(row=table_top, column=1).border = border
        for j, sid in enumerate(ids, start=2):
            c = ws.cell(row=table_top, column=j, value=sid)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        best_id = None
        if 'cop' in selected_params:
            cop_vals = {sid: id_rows[sid].get('cop') for sid in ids if id_rows[sid].get('cop') is not None}
            if cop_vals:
                best_id = max(cop_vals, key=cop_vals.get)

        for i, pkey in enumerate(selected_params, start=table_top + 1):
            ws.cell(row=i, column=1, value=PARAM_META[pkey]['label']).border = border
            ws.cell(row=i, column=1).font = Font(bold=True)
            for j, sid in enumerate(ids, start=2):
                val = id_rows[sid].get(pkey)
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = border
                cell.alignment = center
                if pkey == 'cop' and sid == best_id:
                    cell.fill = best_fill

        # column widths
        ws.column_dimensions['A'].width = 20
        for j in range(2, max(n_data_cols, 4) + 2):
            ws.column_dimensions[get_column_letter(j)].width = 16

        # ---- Embed headline trend charts below the table ----
        chart_row = table_top + len(selected_params) + 3
        col_cursor = 1
        for pkey in HEADLINE_PARAMS:
            png_buf = make_tier_chart_png(tier_block, pkey)
            img = XLImage(png_buf)
            img.width, img.height = 340, 170
            anchor_col = get_column_letter(col_cursor)
            ws.add_image(img, f"{anchor_col}{chart_row}")
            col_cursor += 6

    # Summary sheet
    ws = wb.create_sheet(title="Summary", index=0)
    ws.cell(row=1, column=1, value="Compressor Calorimeter Comparison Summary").font = title_font
    s = payload['summary']
    lines = [
        ("Total Reports Analyzed", s['total_reports']),
        ("Unique Compressor IDs", s['unique_ids']),
        ("RPM Test Tiers", s['rpm_tiers']),
        ("Average COP (all)", s['avg_cop']),
        ("Max COP (all)", s['max_cop']),
        ("Min COP (all)", s['min_cop']),
    ]
    for i, (label, val) in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(rows, selected_params):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                     Spacer, Image as RLImage, PageBreak)

    payload = build_comparison_payload(rows)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=14*mm, rightMargin=14*mm,
                             topMargin=14*mm, bottomMargin=14*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleX', parent=styles['Title'], textColor=colors.HexColor('#1F3864'), fontSize=18)
    sub_style = ParagraphStyle('SubX', parent=styles['Normal'], textColor=colors.HexColor('#444444'), fontSize=10)
    tier_style = ParagraphStyle('TierX', parent=styles['Heading2'], textColor=colors.HexColor('#1F3864'))

    elements = []
    elements.append(Paragraph("Walton Compressor Calorimeter Comparison Report", title_style))
    s = payload['summary']
    elements.append(Paragraph(
        f"Reports analyzed: {s['total_reports']} &nbsp;|&nbsp; Compressor IDs: {s['unique_ids']} "
        f"&nbsp;|&nbsp; RPM tiers: {s['rpm_tiers']} &nbsp;|&nbsp; Avg COP: {s['avg_cop']}",
        sub_style))
    elements.append(Spacer(1, 10))

    for t_idx, tier_block in enumerate(payload['tiers']):
        tier = tier_block['tier']
        ids = tier_block['ids']
        id_rows = tier_block['rows']
        stats = tier_block['stats']

        elements.append(Paragraph(f"{tier} RPM Test Point", tier_style))

        # ---- Headline stats table (range + std dev) ----
        stat_header = ["Metric", "Min", "Max", "Avg", "Std Dev"]
        stat_data = [stat_header]
        for pkey in HEADLINE_PARAMS:
            sdict = stats.get(pkey)
            label = PARAM_META[pkey]['label']
            if sdict:
                stat_data.append([
                    label,
                    PARAM_META[pkey]['fmt'].format(sdict['min']),
                    PARAM_META[pkey]['fmt'].format(sdict['max']),
                    PARAM_META[pkey]['fmt'].format(sdict['avg']),
                    PARAM_META[pkey]['fmt'].format(sdict['std']),
                ])
            else:
                stat_data.append([label, "-", "-", "-", "-"])

        stat_table = Table(stat_data, colWidths=[45*mm, 30*mm, 30*mm, 30*mm, 30*mm])
        stat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F3864')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#EEF3FA')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFBFBF')),
        ]))
        elements.append(stat_table)
        elements.append(Spacer(1, 8))

        # ---- Transposed comparison table: Parameter rows x Compressor ID columns ----
        header = ["Parameter"] + ids
        table_data = [header]
        best_id = None
        if 'cop' in selected_params:
            cop_vals = {sid: id_rows[sid].get('cop') for sid in ids if id_rows[sid].get('cop') is not None}
            if cop_vals:
                best_id = max(cop_vals, key=cop_vals.get)

        for pkey in selected_params:
            row_vals = [PARAM_META[pkey]['label']]
            for sid in ids:
                val = id_rows[sid].get(pkey)
                row_vals.append("-" if val is None else PARAM_META[pkey]['fmt'].format(val))
            table_data.append(row_vals)

        col_w = [42*mm] + [max(22*mm, 180*mm / max(len(ids), 1))] * len(ids)
        t = Table(table_data, repeatRows=1, colWidths=col_w)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F3864')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFBFBF')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F6FC')]),
        ]
        if best_id and best_id in ids and 'cop' in selected_params:
            cop_row_idx = 1 + selected_params.index('cop')
            col_idx = 1 + ids.index(best_id)
            style_cmds.append(('BACKGROUND', (col_idx, cop_row_idx), (col_idx, cop_row_idx), colors.HexColor('#D9EAD3')))
        t.setStyle(TableStyle(style_cmds))
        elements.append(t)
        elements.append(Spacer(1, 10))

        # ---- Headline trend charts ----
        chart_row = []
        for pkey in HEADLINE_PARAMS:
            png_buf = make_tier_chart_png(tier_block, pkey)
            chart_row.append(RLImage(png_buf, width=85*mm, height=42*mm))
        chart_table = Table([chart_row], colWidths=[90*mm, 90*mm, 90*mm])
        chart_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        elements.append(chart_table)

        if t_idx < len(payload['tiers']) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    buf.seek(0)
    return buf


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
